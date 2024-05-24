import os
import shutil
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

class DirectoryManager:
    """A documentar"""
    def __init__(self, origin_directory, destiny_directory):
        self.origin_directory = origin_directory
        self.destiny_directory = destiny_directory
    
    def _list_files(self):
        files = [f for f in os.listdir(self.origin_directory) if os.path.isfile(os.path.join(self.origin_directory, f))]
        full_filenames = [os.path.join(self.origin_directory, f) for f in files]
        return pd.DataFrame({
            'filename': files,
            'full_filename': full_filenames
        })
        
    def _organize_files(self, df):
        for idx, row in df.iterrows():
            if 'pasta' in df.columns and 'subpasta' in df.columns:
                folder_path = os.path.join(self.destiny_directory, row['pasta'], row['subpasta'])
                os.makedirs(folder_path, exist_ok=True)
                file_path = os.path.join(self.origin_directory, row['filename'])
                new_path = os.path.join(folder_path, row['filename'])
                if os.path.exists(file_path):
                    shutil.copy(file_path, new_path)  # Copiar em vez de mover
    
    def execute_organize_files(self, df_folders):
        df = self._list_files()
        df['tabela'] = df['filename'].str.extract('(\d+)')  # Extrai o número
        df['tabela'] = df['tabela'].astype(str)  # Converte para string

        df_folders['tabela'] = df_folders['tabela'].astype(str)  

        if 'tabela' in df_folders.columns:
            df = df.merge(df_folders, on='tabela', how='inner') 
            self._organize_files(df)

    def process_template(self, df, template_path, existing_file_path, tabela):
        template_wb = load_workbook(template_path)
        existing_wb = load_workbook(existing_file_path)
        
        template_sheet = template_wb.active
        new_sheet = existing_wb.create_sheet("Descrição", index=0)  # Nomeia a nova aba
        
        for col in template_sheet.columns:
            for cell in col:
                new_col_letter = get_column_letter(cell.column)
                new_sheet.column_dimensions[new_col_letter].width = template_sheet.column_dimensions[new_col_letter].width
                break  

        for row_index, (index, row) in enumerate(df.iterrows(), start=1):
            for col_index, (key, value) in enumerate(row.items(), start=1):
                # Definindo a célula na nova planilha e copiando o valor
                new_cell = new_sheet.cell(row=row_index, column=col_index)
                new_cell.value = value

                template_cell = template_sheet.cell(row=row_index, column=col_index)
                if template_cell.has_style:
                    new_cell.font = copy(template_cell.font)
                    new_cell.border = copy(template_cell.border)
                    new_cell.fill = copy(template_cell.fill)
                    new_cell.number_format = template_cell.number_format
                    new_cell.alignment = copy(template_cell.alignment)

        existing_wb.save(os.path.join(self.destiny_directory, f'Tabela {tabela}.xlsx'))


# if __name__ == "__main__":

#     output_dir: str = os.path.join(os.path.dirname(__file__), "..", "data", "gold")
#     dm = DirectoryManager(output_dir, None)
#     db = PostgreSQL(schema='datasetpi')
    

#     df = dm._list_files()

#     df['tabela'] = df['filename'].str.extract('(\d+)')  
#     df['tabela'] = df['tabela'].astype(str) 

#     df_tables = db.read_table_columns('sidra_tabelas', columns=['*'], return_type='dataframe')
#     df_tables['id'] = df_tables['id'].astype(str)

#     df = df.merge(df_tables, left_on='tabela', right_on='id', how='inner') 
#     df_final = df[['tabela', 'filename', 'full_filename', 'assunto']]

#     print(df)
#     # self._organize_files(df)
