import pandas as pd
import os

data_path = os.path.join(os.getcwd(), "data")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

def process_template(template_path, df, existing_file_path):
    template_wb = load_workbook(template_path)
    existing_wb = load_workbook(existing_file_path)
    
    template_sheet = template_wb.active
    new_sheet = existing_wb.create_sheet("Nova Descrição", index=0)  # Nomeia a nova aba
    
    for col in template_sheet.columns:
        for cell in col:
            new_col_letter = get_column_letter(cell.column)
            new_sheet.column_dimensions[new_col_letter].width = template_sheet.column_dimensions[new_col_letter].width
            break  

    for row_index, (index, row) in enumerate(df.iterrows(), start=1):
        for col_index, (key, value) in enumerate(row.items(), start=1):
            # Definindo a célula na nova planilha e copiando o valor
            new_cell = new_sheet.cell(row=row_index, column=col_index)
            new_cell.value = value

            template_cell = template_sheet.cell(row=row_index, column=col_index)
            if template_cell.has_style:
                new_cell.font = copy(template_cell.font)
                new_cell.border = copy(template_cell.border)
                new_cell.fill = copy(template_cell.fill)
                new_cell.number_format = template_cell.number_format
                new_cell.alignment = copy(template_cell.alignment)

    existing_wb.save(existing_file_path)

def processar_json(json_data):
    info_geral = {k: v for k, v in json_data.items() if k != 'variaveis'}
    info_geral['Frequência'] = json_data['periodicidade']['frequencia']
    info_geral['Data Inicial'] = json_data['periodicidade']['inicio']
    info_geral['Data Final'] = json_data['periodicidade']['fim']
    
    info_geral['Nível Territorial'] = ', '.join(json_data['nivelTerritorial']['Administrativo'])
    
    del info_geral['periodicidade']
    del info_geral['nivelTerritorial']
    
    # Criando o primeiro DataFrame
    df1 = pd.DataFrame([info_geral])

    info_geral_transposed = [(k, v) for k, v in info_geral.items()]
    info_geral_transposed.append(("", ""))
    info_geral_transposed.append(("Variáveis:", "Unidades:"))

    for var in json_data['variaveis']:
        info_geral_transposed.append((var['nome'], f"{var['unidade']} (Sumarização: {', '.join(var['sumarizacao'])})"))
    
    # Criando o segundo DataFrame transposto com pares de campo e informação
    df2_transposed = pd.DataFrame(info_geral_transposed, columns=['Campo', 'Informação'])

    process_template(f'{data_path}/template.xlsx', df2_transposed, f'{data_path}/tabela_teste.xlsx')
    return df1, df2_transposed

# Usar a função
json_data = {
    'id': 9257,
    'nome': 'Número de estabelecimentos agropecuários com equinos, Número de cabeças de equinos nos estabelecimentos agropecuários, Número de estabelecimentos agropecuários que venderam equinos, Número de cabeças de equinos vendidas nos estabelecimentos agropecuários e Valor da venda de cabeças de equinos nos estabelecimentos agropecuários, por existência de dirigentes de cor ou raça indígena (produtor ou cônjuge) versus agricultura familiar, grupos de atividade econômica e grupos de área total',
    'URL': 'https://sidra.ibge.gov.br/tabela/9257',
    'pesquisa': 'Censo Agropecuário',
    'assunto': 'Pecuária',
    'periodicidade': {'frequencia': 'anual', 'inicio': 2017, 'fim': 2017},
    'nivelTerritorial': {'Administrativo': ['N1', 'N2', 'N6', 'N3'], 'Especial': [], 'IBGE': []},
    'variaveis': [{'id': 42, 'nome': 'Número de estabelecimentos agropecuários com equinos', 'unidade': 'Unidades', 'sumarizacao': ['nivelTerritorial']},
                  {'id': 2196, 'nome': 'Número de cabeças de equinos nos estabelecimentos agropecuários', 'unidade': 'Cabeças', 'sumarizacao': ['nivelTerritorial']},
                  {'id': 10019, 'nome': 'Número de estabelecimentos agropecuários que venderam equinos', 'unidade': 'Unidades', 'sumarizacao': ['nivelTerritorial']},
                  {'id': 10020, 'nome': 'Número de cabeças de equinos vendidas nos estabelecimentos agropecuários', 'unidade': 'Cabeças', 'sumarizacao': ['nivelTerritorial']}]
}

df1, df2 = processar_json(json_data)
print(df1)
print(df2)

# Para exportar para Excel
with pd.ExcelWriter(f'{data_path}/dados_agropecuarios.xlsx') as writer:
    df1.to_excel(writer, sheet_name='Informações Gerais', index=False)
    df2.to_excel(writer, sheet_name='Detalhes Variáveis', index=False)
