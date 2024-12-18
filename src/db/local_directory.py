import os
import shutil
import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class DirectoryManager:
    """Gerencia diretórios locais para organizar e processar arquivos.

    Esta classe fornece métodos para criar diretórios, listar arquivos, 
    organizar arquivos em subpastas e processar arquivos de template.

    Args:
        base_directory (str): O diretório base onde os diretórios "gold", "silver" e "bronze" serão criados. 
            O padrão é um diretório "data" na raiz do projeto.
        origin_directory (str): O diretório de origem dos arquivos a serem organizados.
        destiny_directory (str): O diretório de destino onde os arquivos organizados serão armazenados.

    Attributes:
        base_directory (str): O diretório base onde os diretórios "gold", "silver" e "bronze" serão criados.
        origin_directory (str): O diretório de origem dos arquivos a serem organizados.
        destiny_directory (str): O diretório de destino onde os arquivos organizados serão armazenados.
    """
    def __init__(self, 
                 base_directory: str = os.path.join(os.path.dirname(__file__), "..", "..", "data"), 
                 origin_directory: str = None, 
                 destiny_directory: str = None):
        
        self.base_directory = base_directory
        self.origin_directory = origin_directory
        self.destiny_directory = destiny_directory

    def _create_directories(self):
        """Cria diretórios "gold", "silver" e "bronze" no diretório base se não existirem.

        Returns:
            dict: Um dicionário com os caminhos dos diretórios criados.
        """
        directories = {
            'geral': self.base_directory,
            'gold': os.path.join(self.base_directory, 'gold'),
            'silver': os.path.join(self.base_directory, 'silver'),
            'bronze': os.path.join(self.base_directory, 'bronze')
        }

        for path in directories.values():
            os.makedirs(path, exist_ok=True)
        
        return directories

    def _list_files(self):
        """Lista arquivos no diretório de origem.

        Returns:
            pandas.DataFrame: DataFrame contendo os nomes dos arquivos e seus caminhos completos.
        """
        files = [f for f in os.listdir(self.origin_directory) if os.path.isfile(os.path.join(self.origin_directory, f))]
        full_filenames = [os.path.join(self.origin_directory, f) for f in files]
        
        return pd.DataFrame({
            'filename': files,
            'full_filename': full_filenames
        })

    def _organize_files(self, df):
        """Organiza arquivos em subpastas com base nas informações do DataFrame.

        Args:
            df (pandas.DataFrame): DataFrame contendo informações dos arquivos e as subpastas de destino.
        """
        for idx, row in df.iterrows():
            if 'pasta' in df.columns and 'subpasta' in df.columns:
                folder_path = os.path.join(self.destiny_directory, row['pasta'], row['subpasta'])
                os.makedirs(folder_path, exist_ok=True)
                file_path = os.path.join(self.origin_directory, row['filename'])
                new_path = os.path.join(folder_path, row['filename'])
                if os.path.exists(file_path):
                    shutil.copy(file_path, new_path)  # Copiar em vez de mover

    def execute_organize_files(self, df_folders):
        """Executa a organização dos arquivos com base no DataFrame fornecido.

        Args:
            df_folders (pandas.DataFrame): DataFrame contendo as colunas 'tabela', 'pasta' e 'subpasta' 
                para organizar os arquivos.
        """
        df = self._list_files()
        df['tabela'] = df['filename'].str.extract('(\d+)')  # Extrai o número
        df['tabela'] = df['tabela'].astype(str)  # Converte para string

        df_folders['tabela'] = df_folders['tabela'].astype(str)  

        if 'tabela' in df_folders.columns:
            df = df.merge(df_folders, on='tabela', how='inner') 
            self._organize_files(df)

    def process_template_file(self, df, template_path, existing_file_path, tabela):
        """Processa arquivos de template e adiciona dados de um DataFrame.

        Args:
            df (pandas.DataFrame): DataFrame contendo os dados a serem adicionados ao template.
            template_path (str): Caminho para o arquivo de template.
            existing_file_path (str): Caminho para o arquivo existente onde os dados serão adicionados.
            tabela (str): Nome da tabela para o arquivo de saída.
        """
        template_wb = load_workbook(template_path)
        existing_wb = load_workbook(existing_file_path)
        
        template_sheet = template_wb.active
        new_sheet = existing_wb.create_sheet("Descrição", index=0)  # Nomeia a nova aba
        
        for col in template_sheet.columns:
            for cell in col:
                new_col_letter = get_column_letter(cell.column)
                new_sheet.column_dimensions[new_col_letter].width = template_sheet.column_dimensions[new_col_letter].width
                break  

        for row_index, (index, row) in enumerate(df.iterrows(), start=1):
            for col_index, (key, value) in enumerate(row.items(), start=1):
                new_cell = new_sheet.cell(row=row_index, column=col_index)
                new_cell.value = value

                template_cell = template_sheet.cell(row=row_index, column=col_index)
                if template_cell.has_style:
                    new_cell.font = copy(template_cell.font)
                    new_cell.border = copy(template_cell.border)
                    new_cell.fill = copy(template_cell.fill)
                    new_cell.number_format = template_cell.number_format
                    new_cell.alignment = copy(template_cell.alignment)

        existing_wb.save(os.path.join(self.destiny_directory, f'Tabela {tabela}.xlsx'))
